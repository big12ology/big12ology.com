#!/usr/bin/env python3
"""One-time import: extract the 2025 season from the original Google Sheet xlsx export.

Produces data/seasons/2025.json and data/teams.json from the audited spreadsheet
(Big12 Attendance tab). Kept in the repo for provenance; the 2026 season is
populated by fetch_attendance.py instead.

Usage: python3 scripts/export_2025_from_sheet.py path/to/sheet.xlsx
"""
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent

# Sheet layout: A=team B=stadium C=games D=capacity, then E..AG alternate
# attendance/percent for weeks 0..14. AI/AJ are season totals.
ATT_COLS = [5 + 2 * i for i in range(15)]  # E, G, ..., AG -> weeks 0..14


def main(xlsx_path: str) -> None:
    ws = openpyxl.load_workbook(xlsx_path, data_only=True)["Big12 Attendance"]

    teams = []
    games = []
    for r in range(2, 18):
        team = ws.cell(r, 1).value
        teams.append(
            {
                "team": team,
                "stadium": ws.cell(r, 2).value,
                "capacity": int(ws.cell(r, 4).value),
            }
        )
        for week, c in enumerate(ATT_COLS):
            att = ws.cell(r, c).value
            if att is not None:
                games.append({"team": team, "week": week, "attendance": int(att)})

    games.sort(key=lambda g: (g["week"], g["team"]))

    # NOTE: data/teams.json is hand-curated (per-year capacities with verified
    # sources) and is deliberately NOT written here anymore.
    (ROOT / "data" / "seasons" / "2025.json").write_text(
        json.dumps(
            {
                "season": 2025,
                "source": "Manual entry via Google Sheet (imported)",
                "weekLabels": [f"Week {w}" for w in range(15)],
                "games": games,
            },
            indent=2,
        )
        + "\n"
    )
    print(f"Wrote {len(teams)} teams, {len(games)} games")


if __name__ == "__main__":
    main(sys.argv[1])
